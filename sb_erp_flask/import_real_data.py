"""
sb_erp_flask/data/ 안의 실제 ERP 엑셀(매출자료.xlsx, 외주인건비.xlsx)을
sb_erp_flask/erp_data.db(SQLite)로 적재하는 1회성 스크립트.

사용법: python import_real_data.py
- 여러 번 실행해도 안전함 (매번 기존 데이터를 지우고 다시 적재).
- erp_data.db, data/ 폴더는 .gitignore 처리되어 있어 실제 데이터가
  GitHub에 올라가지 않는다.
"""
import os
import sqlite3
from datetime import datetime

from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
DB_PATH = os.path.join(BASE_DIR, 'erp_data.db')

SALES_FILE = os.path.join(DATA_DIR, '매출자료.xlsx')
OUTSOURCING_FILE = os.path.join(DATA_DIR, '외주인건비.xlsx')


def to_date_str(value):
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    return None


def create_tables(conn):
    conn.execute('''
        CREATE TABLE IF NOT EXISTS sales_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            plant TEXT, customer TEXT, vessel TEXT, product_group TEXT,
            product_code TEXT, product_name TEXT, spec TEXT,
            por_no TEXT, order_no TEXT, order_date TEXT, contract_no TEXT,
            sales_date TEXT, quantity INTEGER, amount INTEGER,
            purchase_flag TEXT, settlement_flag TEXT, category TEXT,
            accounting_flag TEXT, design_team TEXT
        )
    ''')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_sales_date ON sales_records(sales_date)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_sales_customer ON sales_records(customer)')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS outsourcing_costs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            expense_doc_no TEXT, work_team TEXT, year_month TEXT, billing_type TEXT,
            vessel TEXT, product TEXT, item_area TEXT, billing_detail TEXT,
            billing_amount INTEGER, expense_site TEXT, executing_plant TEXT
        )
    ''')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_outsourcing_ym ON outsourcing_costs(year_month)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_outsourcing_team ON outsourcing_costs(work_team)')
    conn.commit()


def import_sales(conn):
    wb = load_workbook(SALES_FILE, read_only=True, data_only=True)
    ws = wb['Sheet1']

    records = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        records.append((
            r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8], r[9],
            to_date_str(r[10]), r[11], to_date_str(r[12]), r[13], r[14],
            r[15], r[16], r[17], r[18], r[19]
        ))

    conn.execute('DELETE FROM sales_records')
    conn.executemany('''
        INSERT INTO sales_records (
            plant, customer, vessel, product_group, product_code, product_name, spec,
            por_no, order_no, order_date, contract_no, sales_date, quantity, amount,
            purchase_flag, settlement_flag, category, accounting_flag, design_team
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', records)
    conn.commit()
    return len(records)


def import_outsourcing(conn):
    wb = load_workbook(OUTSOURCING_FILE, read_only=True, data_only=True)
    ws = wb['1']

    records = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        year_month = str(r[3]) if r[3] is not None else None
        records.append((
            r[1], r[2], year_month, r[4], r[5], r[6], r[7], r[8], r[9], r[10], r[11]
        ))

    conn.execute('DELETE FROM outsourcing_costs')
    conn.executemany('''
        INSERT INTO outsourcing_costs (
            expense_doc_no, work_team, year_month, billing_type, vessel, product,
            item_area, billing_detail, billing_amount, expense_site, executing_plant
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', records)
    conn.commit()
    return len(records)


if __name__ == '__main__':
    if not os.path.exists(SALES_FILE) or not os.path.exists(OUTSOURCING_FILE):
        raise SystemExit(
            f'엑셀 파일을 찾을 수 없습니다. data/ 폴더에 다음 파일이 있어야 합니다:\n'
            f'  {SALES_FILE}\n  {OUTSOURCING_FILE}'
        )

    conn = sqlite3.connect(DB_PATH)
    create_tables(conn)
    sales_count = import_sales(conn)
    outsourcing_count = import_outsourcing(conn)
    conn.close()

    print(f'매출자료 {sales_count:,}건, 외주인건비 {outsourcing_count:,}건 적재 완료 -> {DB_PATH}')
