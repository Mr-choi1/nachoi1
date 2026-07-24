import copy
import csv
import io
import json
import os
import re
import sqlite3
import unicodedata
from collections import Counter
from datetime import datetime, timedelta
from functools import wraps
from urllib.parse import quote

import requests
from dotenv import load_dotenv
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from werkzeug.security import generate_password_hash, check_password_hash

# sb_erp_flask/.env 파일이 있으면 그 안의 값들을 환경변수로 읽어들인다.
load_dotenv()

app = Flask(__name__)
# 운영 배포 시 반드시 SECRET_KEY 환경변수로 교체할 것 (세션 쿠키 서명에 사용됨)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

# 세션 쿠키 보안 설정. SECURE는 HTTPS 환경(FLASK_ENV=production)에서만 켠다 —
# 로컬 http 개발 환경에서 켜두면 쿠키가 전송되지 않아 로그인이 막힌다.
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('FLASK_ENV') == 'production'

LOGIN_MAX_ATTEMPTS = 5
LOGIN_LOCKOUT_MINUTES = 15


@app.after_request
def set_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['Referrer-Policy'] = 'no-referrer'
    # index.html이 chart.js를 CDN에서 불러오고 스타일/스크립트가 인라인으로
    # 작성되어 있어 script-src/style-src에 unsafe-inline이 불가피하다.
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' https://cdn.jsdelivr.net 'unsafe-inline'; "
        "style-src 'self' 'unsafe-inline'; "
        "img-src 'self' data:;"
    )
    if request.is_secure:
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    return response


# ---------- 사용자 DB (SQLite) ----------
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'users.db')

# 부서별로 조회 가능한 데이터 카테고리 (RBAC). role이 admin이면 전체 조회 가능.
DEPARTMENT_CATEGORY_ACCESS = {
    '관리부': {'financial', 'production', 'purchase', 'quality', 'quality_cause', 'hr', 'comprehensive'},
    '영업부': {'financial', 'purchase', 'comprehensive'},
    '생산부': {'production', 'quality', 'quality_cause', 'comprehensive'},
    '품질부': {'quality', 'quality_cause', 'production', 'comprehensive'},
    '기술부': {'production', 'purchase', 'comprehensive'},
}

ALL_CATEGORIES = {'financial', 'production', 'purchase', 'quality', 'quality_cause', 'hr', 'comprehensive'}

# 특정 불량 유형이 전체 불량의 이 비율(%)을 넘으면 관련 부서에 알림 발송
QUALITY_ALERT_THRESHOLD = 35


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            department TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'staff'
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS notifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            department TEXT NOT NULL,
            message TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime')),
            is_read INTEGER NOT NULL DEFAULT 0,
            resolution_comment TEXT
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS conversations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            title TEXT NOT NULL,
            is_favorite INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime')),
            updated_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime'))
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS query_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            conversation_id INTEGER,
            query TEXT NOT NULL,
            response_type TEXT NOT NULL,
            response_message TEXT NOT NULL,
            response_data TEXT,
            created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime'))
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS login_attempts (
            username TEXT PRIMARY KEY,
            failed_count INTEGER NOT NULL DEFAULT 0,
            locked_until TEXT
        )
    ''')
    conn.commit()

    existing = conn.execute('SELECT COUNT(*) AS cnt FROM users').fetchone()['cnt']
    if existing == 0:
        # 최초 실행 시에만 데모 계정 시드. 운영 전 반드시 비밀번호 변경 필요.
        seed_users = [
            ('admin', 'admin1234', '관리부', 'admin'),
            ('sales1', 'sales1234', '영업부', 'staff'),
            ('prod1', 'prod1234', '생산부', 'staff'),
            ('qc1', 'qc1234', '품질부', 'staff'),
            ('tech1', 'tech1234', '기술부', 'staff'),
        ]
        for username, pw, dept, role in seed_users:
            conn.execute(
                'INSERT INTO users (username, password_hash, department, role) VALUES (?, ?, ?, ?)',
                (username, generate_password_hash(pw), dept, role)
            )
        conn.commit()
    conn.close()


def get_allowed_categories(department, role):
    if role == 'admin':
        return set(ALL_CATEGORIES)
    return set(DEPARTMENT_CATEGORY_ACCESS.get(department, set()))


# ---------- 로그인 브루트포스 방지 ----------
def is_locked_out(username):
    conn = get_db()
    row = conn.execute('SELECT locked_until FROM login_attempts WHERE username = ?', (username,)).fetchone()
    conn.close()
    if row and row['locked_until']:
        locked_until = datetime.strptime(row['locked_until'], '%Y-%m-%d %H:%M:%S')
        if datetime.now() < locked_until:
            return locked_until
    return None


def record_failed_login(username):
    conn = get_db()
    row = conn.execute('SELECT failed_count FROM login_attempts WHERE username = ?', (username,)).fetchone()
    failed_count = (row['failed_count'] if row else 0) + 1
    locked_until = None
    if failed_count >= LOGIN_MAX_ATTEMPTS:
        locked_until = (datetime.now() + timedelta(minutes=LOGIN_LOCKOUT_MINUTES)).strftime('%Y-%m-%d %H:%M:%S')
        failed_count = 0

    conn.execute('''
        INSERT INTO login_attempts (username, failed_count, locked_until) VALUES (?, ?, ?)
        ON CONFLICT(username) DO UPDATE SET failed_count = excluded.failed_count, locked_until = excluded.locked_until
    ''', (username, failed_count, locked_until))
    conn.commit()
    conn.close()
    return locked_until is not None


def reset_failed_login(username):
    conn = get_db()
    conn.execute('DELETE FROM login_attempts WHERE username = ?', (username,))
    conn.commit()
    conn.close()


# ---------- 임계치 알림 ----------
def create_notification(department, message):
    conn = get_db()
    conn.execute(
        'INSERT INTO notifications (department, message) VALUES (?, ?)',
        (department, message)
    )
    conn.commit()
    conn.close()


def notify_if_new(department, message):
    # 같은 경고가 미확인 상태로 이미 있으면 중복 알림을 만들지 않는다.
    conn = get_db()
    existing = conn.execute(
        'SELECT id FROM notifications WHERE department = ? AND message = ? AND is_read = 0',
        (department, message)
    ).fetchone()
    conn.close()
    if not existing:
        create_notification(department, message)


def get_notifications(department, role, unread_only=False):
    conn = get_db()
    if role == 'admin':
        query = 'SELECT * FROM notifications'
        params = ()
    else:
        query = 'SELECT * FROM notifications WHERE department = ?'
        params = (department,)
    if unread_only:
        query += (' AND' if 'WHERE' in query else ' WHERE') + ' is_read = 0'
    query += ' ORDER BY created_at DESC LIMIT 50'
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return [dict(row) for row in rows]


# ---------- 대화 세션 (Claude/ChatGPT 스타일 대화 기록) ----------
def create_conversation(user_id, title):
    conn = get_db()
    cur = conn.execute('INSERT INTO conversations (user_id, title) VALUES (?, ?)', (user_id, title[:60]))
    conn.commit()
    conversation_id = cur.lastrowid
    conn.close()
    return conversation_id


def touch_conversation(conversation_id):
    conn = get_db()
    conn.execute("UPDATE conversations SET updated_at = datetime('now', 'localtime') WHERE id = ?", (conversation_id,))
    conn.commit()
    conn.close()


def get_conversations(user_id, limit=50):
    conn = get_db()
    rows = conn.execute(
        'SELECT * FROM conversations WHERE user_id = ? ORDER BY updated_at DESC LIMIT ?',
        (user_id, limit)
    ).fetchall()
    conn.close()
    return [dict(row) for row in rows]


def get_conversation_messages(conversation_id, user_id):
    conn = get_db()
    rows = conn.execute(
        'SELECT * FROM query_history WHERE conversation_id = ? AND user_id = ? ORDER BY id ASC',
        (conversation_id, user_id)
    ).fetchall()
    conn.close()

    messages = []
    for row in rows:
        item = dict(row)
        if item.get('response_data'):
            try:
                item['response_data'] = json.loads(item['response_data'])
            except ValueError:
                item['response_data'] = None
        messages.append(item)
    return messages


def get_last_domain_in_conversation(conversation_id, user_id):
    if not conversation_id:
        return None
    conn = get_db()
    row = conn.execute('''
        SELECT response_type FROM query_history
        WHERE conversation_id = ? AND user_id = ?
        AND response_type IN ('financial', 'production', 'purchase', 'quality', 'quality_cause', 'hr')
        ORDER BY id DESC LIMIT 1
    ''', (conversation_id, user_id)).fetchone()
    conn.close()
    return row['response_type'] if row else None


def log_query_history(user_id, conversation_id, query, response_type, response_message, response_data=None):
    conn = get_db()
    conn.execute(
        'INSERT INTO query_history (user_id, conversation_id, query, response_type, response_message, response_data) '
        'VALUES (?, ?, ?, ?, ?, ?)',
        (
            user_id, conversation_id, query, response_type, response_message,
            json.dumps(response_data, ensure_ascii=False) if response_data is not None else None
        )
    )
    conn.commit()
    conn.close()


def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if 'user_id' not in session:
            if request.path.startswith('/api/'):
                return jsonify({'error': '로그인이 필요합니다.'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return wrapper


@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        locked_until = is_locked_out(username)
        if locked_until:
            error = f'로그인 시도가 너무 많아 {locked_until.strftime("%H:%M")}까지 잠겼습니다. 잠시 후 다시 시도해주세요.'
        else:
            conn = get_db()
            user = conn.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
            conn.close()

            if user and check_password_hash(user['password_hash'], password):
                reset_failed_login(username)
                session['user_id'] = user['id']
                session['username'] = user['username']
                session['department'] = user['department']
                session['role'] = user['role']
                return redirect(url_for('index'))

            just_locked = record_failed_login(username)
            error = (
                f'로그인 시도가 너무 많아 {LOGIN_LOCKOUT_MINUTES}분간 잠겼습니다.'
                if just_locked else '아이디 또는 비밀번호가 올바르지 않습니다.'
            )

    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ---------- 포텐스닷(potens.ai) LLM 연동 ----------
# 키는 코드에 직접 넣지 말고 환경변수 POTENS_API_KEY 로 설정합니다.
POTENS_API_URL = 'https://ai.potens.ai/api/chat'
POTENS_MODEL = 'claude-4-6-sonnet'


def call_potens_ai(prompt):
    api_key = os.environ.get('POTENS_API_KEY')
    if not api_key:
        raise RuntimeError('POTENS_API_KEY 환경변수가 설정되어 있지 않습니다.')

    response = requests.post(
        POTENS_API_URL,
        json={'prompt': prompt, 'model': POTENS_MODEL},
        headers={'Authorization': f'Bearer {api_key}'},
        timeout=45
    )
    response.raise_for_status()
    return response.json().get('message', '')


# ---------- RAG: 사내 문서 검색 ----------
# 벡터DB 없이 키워드 겹침 기반의 경량 검색으로, knowledge/ 폴더의 문서 중
# 질문과 가장 관련 있는 문서를 찾아 LLM 프롬프트에 근거로 첨부한다.
KNOWLEDGE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'knowledge')


def load_knowledge_documents():
    docs = []
    if not os.path.isdir(KNOWLEDGE_DIR):
        return docs
    for fname in sorted(os.listdir(KNOWLEDGE_DIR)):
        if fname.endswith('.md') or fname.endswith('.txt'):
            path = os.path.join(KNOWLEDGE_DIR, fname)
            with open(path, encoding='utf-8') as f:
                docs.append({'title': fname.rsplit('.', 1)[0], 'content': f.read()})
    return docs


def char_bigrams(text):
    # 한글은 조사(는/은/가/을 등)가 단어에 바로 붙기 때문에 단어 단위 매칭은
    # "휴가는"과 "휴가"를 다른 단어로 취급해 놓치기 쉽다. 음절 2-gram(바이그램)
    # 기반으로 비교하면 조사가 붙어도 어간 부분이 겹쳐서 매칭된다.
    cleaned = re.sub(r'[^가-힣a-zA-Z0-9]', '', text.lower())
    return Counter(cleaned[i:i + 2] for i in range(len(cleaned) - 1))


def search_knowledge(query, top_k=2):
    query_grams = char_bigrams(query)
    total = sum(query_grams.values())
    if total == 0:
        return []

    scored = []
    for doc in load_knowledge_documents():
        doc_grams = char_bigrams(doc['title'] + doc['content'])
        overlap = sum((query_grams & doc_grams).values())
        if overlap:
            scored.append((overlap / total, doc))

    scored = [(score, doc) for score, doc in scored if score >= 0.15]
    if not scored:
        return []

    scored.sort(key=lambda x: x[0], reverse=True)
    top_score = scored[0][0]
    # 1등과 점수 차이가 큰(우연히 겹친) 문서는 노이즈이므로 함께 반환하지 않는다.
    return [doc for score, doc in scored[:top_k] if score >= top_score * 0.5]


# ---------- 자주 묻는 질문 자동 매뉴얼(FAQ) 생성 ----------
# 반복적으로 들어온 질문·답변을 knowledge/ 폴더에 문서로 누적해, 다음부터는
# RAG 검색으로도 같은 질문에 답할 수 있게 한다 (스텝하우의 "자동 매뉴얼 누적" 개념).
def generate_faq_document(min_count=2, top_n=15):
    conn = get_db()
    rows = conn.execute('''
        SELECT query, response_message, COUNT(*) AS cnt
        FROM query_history
        WHERE response_type != 'text'
        GROUP BY query
        HAVING cnt >= ?
        ORDER BY cnt DESC
        LIMIT ?
    ''', (min_count, top_n)).fetchall()
    conn.close()

    if not rows:
        return None

    lines = [
        '# 자주 묻는 질문 (자동 생성)',
        '',
        '이 문서는 챗봇 조회 이력을 기반으로 자동 생성되었습니다. 반복적으로 들어온 질문과',
        '가장 최근 응답을 정리한 것으로, 실제 최신 수치는 챗봇에 직접 질의해 확인하세요.',
        ''
    ]
    for row in rows:
        lines.append(f"## {row['query']}")
        lines.append(f"(누적 {row['cnt']}회 질의)")
        lines.append('')
        lines.append(row['response_message'])
        lines.append('')

    path = os.path.join(KNOWLEDGE_DIR, '자동생성_FAQ.md')
    os.makedirs(KNOWLEDGE_DIR, exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    return path


# ---------- 부서별 ERP 사용 가이드 자동 생성 ----------
# 스텝하우의 "부서/직무별 매뉴얼" 컨셉을, 화면 캡처 대신 부서 조회 권한 +
# 사내 문서를 조합한 텍스트 가이드로 구현. 부서마다 조회 가능한 항목이 달라
# 자동으로 각자 다른 내용의 가이드가 만들어진다.
EXAMPLE_QUERIES_BY_CATEGORY = {
    'financial': ['3월과 4월 매출 비교해줘', '최근 매출 추이 알려줘', '이익률이 가장 높은 달은?'],
    'production': ['4월 생산량 알려줘', '생산량이 가장 많은 제품은?', 'A-100 모델 생산 현황'],
    'purchase': ['이번 달 구매 발주 현황 알려줘', '완료된 발주 목록 보여줘', '글로벌부품 발주 내역 알려줘'],
    'quality': ['불량의 주요 원인은?', '품질 합격률 알려줘', '외관 불량 현황 보여줘'],
    'hr': ['부서별 출근율 알려줘', '우리 부서 현황 알려줘', '이번 달 입사/퇴사 현황 알려줘'],
    'comprehensive': ['전체 경영 현황 요약해줘']
}


def build_department_guide(department, allowed_categories):
    if not allowed_categories:
        return None

    lines = [
        f'# {department} ERP 사용 가이드 (자동 생성)',
        '',
        f'이 문서는 {department}의 조회 권한을 기준으로 자동 생성되었습니다. '
        '아래 예시처럼 챗봇에 자유롭게 질문하면 됩니다.',
        ''
    ]

    lines.append('## 조회 가능한 데이터 및 예시 질문')
    lines.append('')
    excel_rows = []
    for category in sorted(allowed_categories):
        # quality_cause는 별도 업무 영역이 아니라 '품질' 질문 중 원인분석 표현을 쓸 때
        # 자동으로 걸리는 하위 개념이라, 가이드에서는 품질 섹션 하나로 충분하다.
        if category == 'quality_cause':
            continue
        label = DOMAIN_LABELS.get(category, '종합 현황' if category == 'comprehensive' else category)
        menu = MENU_GUIDE.get(category, '')
        heading = f'### {label}' + (f' ({menu})' if menu else '')
        lines.append(heading)
        for q in EXAMPLE_QUERIES_BY_CATEGORY.get(category, []):
            lines.append(f'- "{q}"')
            excel_rows.append({'분류': label, '관련메뉴': menu, '예시질문': q})
        lines.append('')

    manual_docs = [
        d for d in load_knowledge_documents()
        if not d['title'].startswith('자동생성') and not d['title'].endswith('_가이드')
    ]
    if manual_docs:
        lines.append('## 참고 가능한 사내 규정 문서')
        lines.append('아래 문서 관련 내용은 규정을 그대로 질문해도 답변받을 수 있습니다.')
        for d in manual_docs:
            lines.append(f"- {d['title']}")
        lines.append('')

    return {'markdown': '\n'.join(lines), 'rows': excel_rows}


def save_department_guide_to_knowledge(department, markdown_text):
    path = os.path.join(KNOWLEDGE_DIR, f'{department}_가이드.md')
    os.makedirs(KNOWLEDGE_DIR, exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        f.write(markdown_text)
    return path


# ---------- 오타 보정 (레벤슈타인 유사도) ----------
# 각 카테고리 키워드 목록 자체가 유사어 사전 역할을 하고,
# 여기에 오타까지 허용하기 위해 단어 단위 유사도 매칭을 함께 사용한다.
def calculate_similarity(a, b):
    len1, len2 = len(a), len(b)
    if len1 == 0 or len2 == 0:
        return 1.0 if len1 == len2 else 0.0

    matrix = [[0] * (len2 + 1) for _ in range(len1 + 1)]
    for i in range(len1 + 1):
        matrix[i][0] = i
    for j in range(len2 + 1):
        matrix[0][j] = j

    for i in range(1, len1 + 1):
        for j in range(1, len2 + 1):
            if a[i - 1] == b[j - 1]:
                matrix[i][j] = matrix[i - 1][j - 1]
            else:
                matrix[i][j] = min(
                    matrix[i - 1][j - 1] + 1,
                    matrix[i][j - 1] + 1,
                    matrix[i - 1][j] + 1
                )

    distance = matrix[len1][len2]
    return 1 - distance / max(len1, len2)


# 2음절 키워드는 글자 하나만 달라도 유사도가 0.5까지 떨어져서, 느슨한 기준을
# 쓰면 "인사"~"검사"처럼 전혀 다른 단어까지 오매칭된다. 그래서 2음절 키워드는
# 실사용에서 자주 보이는 오타만 사전에 등록해 명시적으로 허용하고, 4음절
# 이상의 긴 단어에 한해서만 레벤슈타인 유사도 매칭을 함께 사용한다.
TYPO_VARIANTS = {
    '매출': ['메출', '매촐', '매춀'],
    '생산': ['셍산', '생상', '셍상'],
    '불량': ['불냥', '불럥'],
    '직원': ['즉원'],
    '재무': ['재부', '제무'],
    '구매': ['구메'],
    '품질': ['품칠', '풍질'],
    '인사': ['인싸'],
    '부서': ['부셔'],
    '근태': ['근테'],
}


def contains_keyword(query, keywords):
    for keyword in keywords:
        if keyword in query:
            return True
        for variant in TYPO_VARIANTS.get(keyword, []):
            if variant in query:
                return True
        if len(keyword) >= 4:
            for word in query.split():
                if len(word) >= 4 and calculate_similarity(word, keyword) >= 0.75:
                    return True
    return False


# ---------- 웹 앱 실행 ----------
@app.route('/')
@login_required
def index():
    allowed = get_allowed_categories(session['department'], session['role'])
    return render_template(
        'index.html',
        username=session['username'],
        department=session['department'],
        role=session['role'],
        allowed_categories=allowed
    )


# ---------- AI 챗봇 응답 처리 ----------
@app.route('/api/query', methods=['POST'])
@login_required
def process_ai_query_route():
    payload = request.json or {}
    user_query = payload.get('query', '')
    conversation_id = payload.get('conversation_id')

    allowed = get_allowed_categories(session['department'], session['role'])
    last_domain = get_last_domain_in_conversation(conversation_id, session['user_id'])
    result = process_ai_query(user_query, allowed, session['department'], last_domain)

    if user_query.strip():
        if not conversation_id:
            conversation_id = create_conversation(session['user_id'], user_query.strip())
        else:
            touch_conversation(conversation_id)
        log_query_history(
            session['user_id'], conversation_id, user_query,
            result.get('type', 'text'), result.get('message', ''), result.get('data')
        )

    result['conversation_id'] = conversation_id
    return jsonify(result)


def process_ai_query(user_query, allowed_categories=None, department=None, last_domain=None):
    # 브라우저/IME에 따라 한글이 조합형(NFD)으로 넘어오면 '메출' 같은 문자열이
    # 우리 코드의 조합완성형(NFC) 키워드와 바이트 단위로 달라 매칭이 실패할 수 있다.
    # 항상 NFC로 정규화해서 비교한다.
    user_query = unicodedata.normalize('NFC', user_query)
    query = user_query.lower()

    period = extract_period(query)
    is_comparison = contains_keyword(query, ['비교', '차이', '대비'])
    specific_item = extract_specific_item(query)

    # 조직도 연동: "우리 부서/우리팀" 질문은 로그인한 사용자의 소속 부서로 자동 인식
    if specific_item is None and department and contains_keyword(query, ['우리 부서', '우리부서', '우리팀', '우리 팀', '저희 부서', '저희팀']):
        specific_item = department

    stats_type = extract_stats_type(query)
    is_cause_analysis = contains_keyword(query, ['원인', '이유', '왜'])
    # "어떻게 해야 돼", "지금 뭘 확인해야 돼", "우선순위대로 정리해줘" 같은 질문은
    # 결과를 산문이 아니라 번호 매긴 체크리스트로 정리해서 답하도록 유도한다.
    # (RAG 문서 기반이면 절차 안내로, ERP 데이터 기반이면 이슈 우선순위 체크리스트로 쓰인다)
    is_checklist_request = contains_keyword(
        query, ['어떻게', '방법', '절차', '체크리스트', '준비물', '하려면', '순서', '확인해야', '확인할', '우선순위']
    )

    result = None

    if contains_keyword(query, ['재무', '매출', '수익', '손익', '이익']):
        result = handle_financial_query(query, period, is_comparison, specific_item, stats_type)
    elif contains_keyword(query, ['생산', '제품', '생산량', '공장']):
        result = handle_production_query(query, period, is_comparison, specific_item, stats_type)
    elif contains_keyword(query, ['구매', '발주', '자재', '공급']):
        result = handle_purchase_query(query, period, is_comparison, specific_item, stats_type)
    elif contains_keyword(query, ['품질', '불량', '검사', '합격']):
        result = handle_quality_query(query, period, is_comparison, specific_item, stats_type, is_cause_analysis)
    elif contains_keyword(query, ['인사', '직원', '사원', '근태', '부서']):
        result = handle_hr_query(query, period, is_comparison, specific_item, stats_type)
    elif contains_keyword(query, ['전체', '모든', '종합']):
        result = handle_comprehensive_query(query, allowed_categories)
    elif is_checklist_request and last_domain:
        # "얘네 지금 뭐부터 확인해야 돼?" 처럼 도메인 키워드 없이 직전 대화의
        # 주제(last_domain)를 이어받아 체크리스트를 요청하는 경우.
        domain_data = get_domain_data(last_domain)
        if domain_data is not None:
            result = {'type': last_domain, 'data': domain_data, 'message': '', 'query': query}

    if result:
        if allowed_categories is not None and result.get('type') not in allowed_categories:
            return {
                'type': 'text',
                'data': None,
                'message': '🔒 소속 부서 권한으로는 조회할 수 없는 데이터입니다. 담당 부서 또는 관리자에게 문의해주세요.'
            }

        # ERP 데이터를 근거로 지금 확인/조치해야 할 사항을 우선순위 체크리스트로 정리
        if is_checklist_request and result.get('type') in ('financial', 'production', 'purchase', 'quality', 'quality_cause', 'hr'):
            checklist = generate_issue_checklist(result['type'], result.get('data'))
            if checklist:
                result['message'] = checklist
            elif not result['message']:
                result['message'] = f"{DOMAIN_LABELS.get(result['type'], result['type'])} 현황을 조회했습니다."

        return enrich_structured_response(result)

    # 정형 카테고리 키워드에 걸리지 않은 질문은 사내 문서(RAG)에서 근거를 찾은 경우에만
    # AI가 답변한다. 사내 데이터/문서에서 찾을 수 없는 내용은 일반 지식으로 답하지 않고
    # 아래 안내 메시지로 넘어간다 (일반 상식/외부지식 자유 응답은 제공하지 않음).
    knowledge_hits = search_knowledge(user_query)
    if knowledge_hits:
        try:
            context_text = '\n\n'.join(
                f"[{doc['title']}]\n{doc['content']}" for doc in knowledge_hits
            )
            if is_checklist_request:
                instruction = (
                    '다음은 사내 규정/매뉴얼 문서입니다. 이 내용을 근거로, 사용자가 지금 처리해야 할 업무를 '
                    '번호를 매긴 체크리스트 형태로(1. 2. 3. ...) 정리해서 답변하세요. 각 단계는 한 줄로 '
                    '간결하게 쓰고, 필요하면 단계별로 담당 부서나 기한도 함께 표시하세요. '
                    '문서에 없는 내용은 추측하지 말고 모른다고 답하세요.'
                )
            else:
                instruction = (
                    '다음은 사내 규정/매뉴얼 문서입니다. 이 내용을 근거로 사용자 질문에 답변하세요. '
                    '문서에 없는 내용은 추측하지 말고 모른다고 답하세요.'
                )
            prompt = f'{instruction}\n\n{context_text}\n\n사용자 질문: {user_query}'
            ai_answer = call_potens_ai(prompt)
            sources = ', '.join(doc['title'] for doc in knowledge_hits)
            return {
                'type': 'text',
                'data': None,
                'message': (
                    f'📚 {ai_answer}\n\n'
                    f'※ 참고 문서: {sources}'
                )
            }
        except Exception as e:
            print(f'Potens AI 호출 실패: {e}')

    suggestions = get_suggestions(query)
    suggestion_text = ''
    if suggestions:
        suggestion_text = (
            '혹시 이런 질문을 하신 건가요?\n' +
            '\n'.join(f'- "{s}"' for s in suggestions) +
            '\n\n'
        )

    return {
        'type': 'text',
        'data': None,
        'message': (
            '질문을 더 구체적으로 해주시면 정확한 데이터를 찾아드리겠습니다.\n\n' +
            suggestion_text +
            '💡 질문 예시:\n'
            '- "3월과 4월 매출 비교해줘"\n'
            '- "4월 생산량 알려줘"\n'
            '- "불량의 주요 원인은?"\n'
            '- "생산부 출근율 보여줘"\n'
            '- "A-100 모델 생산 현황"'
        )
    }


# ---------- 질문 추천 ----------
def get_suggestions(query):
    templates = [
        '3월과 4월 매출 비교해줘',
        '4월 생산량 알려줘',
        '불량의 주요 원인은?',
        '생산부 출근율 보여줘',
        'A-100 모델 생산 현황',
        '이번 달 구매 발주 현황 알려줘',
        '부서별 인원 현황 보여줘',
        '전체 경영 현황 요약해줘'
    ]

    words = [w for w in query.split() if len(w) >= 2]
    suggestions = []

    for template in templates:
        if any(w in template for w in words):
            suggestions.append(template)
        if len(suggestions) >= 3:
            break

    return suggestions


# ---------- 응답별 신뢰도 표시 + 관련 메뉴 안내 ----------
MENU_GUIDE = {
    'financial': '재무관리 > 손익현황',
    'production': '생산관리 > 생산실적',
    'purchase': '구매관리 > 발주현황',
    'quality': '품질관리 > 검사현황',
    'quality_cause': '품질관리 > 불량원인분석',
    'hr': '인사관리 > 근태/부서현황',
    'comprehensive': '경영현황판 > 전체 요약'
}


def enrich_structured_response(result):
    footer = '✅ 사내 DB 기반 확정 데이터입니다.'
    menu = MENU_GUIDE.get(result.get('type'))
    if menu:
        footer += f'\n📁 관련 메뉴: {menu}'

    result['message'] = f"{result['message']}\n\n{footer}"
    return result


# ---------- 기간 추출 ----------
def extract_period(query):
    months = ['1월', '2월', '3월', '4월', '5월', '6월', '7월', '8월', '9월', '10월', '11월', '12월']
    found_months = [m for m in months if m in query]

    if contains_keyword(query, ['최근', '이번주', '이번달']):
        return {'type': 'recent', 'value': None}
    if contains_keyword(query, ['지난', '전']):
        return {'type': 'past', 'value': None}
    if found_months:
        return {'type': 'specific', 'value': found_months}

    return {'type': 'all', 'value': None}


# ---------- 특정 항목 추출 ----------
def extract_specific_item(query):
    if 'a-100' in query or 'a100' in query:
        return 'A-100 모델'
    if 'b-200' in query or 'b200' in query:
        return 'B-200 모델'
    if 'c-300' in query or 'c300' in query:
        return 'C-300 모델'
    if 'd-400' in query or 'd400' in query:
        return 'D-400 모델'

    if '생산부' in query:
        return '생산부'
    if '영업부' in query:
        return '영업부'
    if '기술부' in query:
        return '기술부'
    if '관리부' in query:
        return '관리부'
    if '품질부' in query:
        return '품질부'

    if '외관' in query:
        return '외관 불량'
    if '치수' in query:
        return '치수 불량'
    if '기능' in query:
        return '기능 불량'
    if '포장' in query:
        return '포장 불량'

    if '대한소재' in query:
        return '대한소재'
    if '글로벌부품' in query:
        return '글로벌부품'
    if '한국화학' in query:
        return '한국화학'
    if '프리미엄자재' in query:
        return '프리미엄자재'
    if '스마트부품' in query:
        return '스마트부품상사'

    return None


# ---------- 통계 유형 추출 ----------
def extract_stats_type(query):
    if contains_keyword(query, ['평균']):
        return 'average'
    if contains_keyword(query, ['최대', '가장 높', '제일 높', '가장 많']):
        return 'max'
    if contains_keyword(query, ['최소', '가장 낮', '제일 낮', '가장 적']):
        return 'min'
    if contains_keyword(query, ['합계', '총']):
        return 'sum'
    if contains_keyword(query, ['증가', '상승']):
        return 'increase'
    if contains_keyword(query, ['감소', '하락']):
        return 'decrease'
    if contains_keyword(query, ['추세', '트렌드']):
        return 'trend'

    return None


# ---------- 재무 질문 처리 ----------
def handle_financial_query(query, period, is_comparison, specific_item, stats_type):
    data = get_financial_data()
    message = ''
    processed_data = copy.deepcopy(data)

    period_label = None
    if period['type'] == 'specific' and period['value']:
        processed_data['monthly'] = [m for m in data['monthly'] if m['month'] in period['value']]
        period_label = ', '.join(period['value'])
    elif period['type'] == 'recent':
        processed_data['monthly'] = data['monthly'][-3:]
        period_label = f"최근 {len(processed_data['monthly'])}개월(" + ', '.join(m['month'] for m in processed_data['monthly']) + ")"
    elif period['type'] == 'past' and len(data['monthly']) >= 2:
        processed_data['monthly'] = [data['monthly'][-2]]
        period_label = f"지난달({processed_data['monthly'][0]['month']})"

    if period_label:
        message += f"📅 {period_label} 재무 데이터:\n\n"

        for m in processed_data['monthly']:
            message += f"[{m['month']}]\n"
            message += f"💰 매출: {format_number(m['revenue'])}원\n"
            message += f"💸 비용: {format_number(m['expense'])}원\n"
            message += f"📈 이익: {format_number(m['profit'])}원\n"
            message += f"📊 수익률: {(m['profit'] / m['revenue'] * 100):.1f}%\n\n"

    if is_comparison and len(processed_data['monthly']) >= 2:
        first = processed_data['monthly'][0]
        last = processed_data['monthly'][-1]
        revenue_diff = last['revenue'] - first['revenue']
        profit_diff = last['profit'] - first['profit']

        message += f"\n📊 비교 분석 ({first['month']} vs {last['month']}):\n"
        message += f"- 매출 변화: {format_number(abs(revenue_diff))}원 {'📈 증가' if revenue_diff > 0 else '📉 감소'} ({(revenue_diff / first['revenue'] * 100):.1f}%)\n"
        message += f"- 이익 변화: {format_number(abs(profit_diff))}원 {'📈 증가' if profit_diff > 0 else '📉 감소'} ({(profit_diff / first['profit'] * 100):.1f}%)\n"

        processed_data['comparison'] = {
            'revenueDiff': revenue_diff,
            'profitDiff': profit_diff,
            'revenuePercent': round(revenue_diff / first['revenue'] * 100, 2),
            'profitPercent': round(profit_diff / first['profit'] * 100, 2)
        }

    if stats_type:
        revenues = [m['revenue'] for m in processed_data['monthly']]
        profits = [m['profit'] for m in processed_data['monthly']]

        if stats_type == 'average':
            avg_revenue = sum(revenues) / len(revenues)
            avg_profit = sum(profits) / len(profits)
            message += f"\n📈 평균 데이터:\n"
            message += f"- 평균 매출: {format_number(round(avg_revenue))}원\n"
            message += f"- 평균 이익: {format_number(round(avg_profit))}원\n"
            processed_data['stats'] = {'avgRevenue': avg_revenue, 'avgProfit': avg_profit}

        elif stats_type == 'max':
            max_revenue = max(revenues)
            max_revenue_month = next(m for m in processed_data['monthly'] if m['revenue'] == max_revenue)
            message += f"\n🔝 최대 매출:\n"
            message += f"- {max_revenue_month['month']}: {format_number(max_revenue)}원\n"
            processed_data['highlight'] = max_revenue_month

        elif stats_type == 'min':
            min_revenue = min(revenues)
            min_revenue_month = next(m for m in processed_data['monthly'] if m['revenue'] == min_revenue)
            message += f"\n📉 최소 매출:\n"
            message += f"- {min_revenue_month['month']}: {format_number(min_revenue)}원\n"
            processed_data['highlight'] = min_revenue_month

        elif stats_type == 'sum':
            total_revenue = sum(revenues)
            total_profit = sum(profits)
            message += f"\n💰 총합:\n"
            message += f"- 총 매출: {format_number(total_revenue)}원\n"
            message += f"- 총 이익: {format_number(total_profit)}원\n"

        elif stats_type == 'trend':
            is_increasing = revenues[-1] > revenues[0]
            message += f"\n📊 추세 분석:\n"
            message += f"- 전반적으로 {'📈 증가' if is_increasing else '📉 감소'} 추세입니다.\n"

    if not message:
        message = '재무 현황을 조회했습니다.'

    return {
        'type': 'financial',
        'data': processed_data,
        'message': message,
        'query': query
    }


# ---------- 생산 질문 처리 ----------
def handle_production_query(query, period, is_comparison, specific_item, stats_type):
    data = get_production_data()
    message = ''
    processed_data = copy.deepcopy(data)

    period_label = None
    if period['type'] == 'specific' and period['value']:
        processed_data['monthly'] = [m for m in data['monthly'] if m['month'] in period['value']]
        period_label = ', '.join(period['value'])
    elif period['type'] == 'recent':
        processed_data['monthly'] = data['monthly'][-3:]
        period_label = f"최근 {len(processed_data['monthly'])}개월(" + ', '.join(m['month'] for m in processed_data['monthly']) + ")"
    elif period['type'] == 'past' and len(data['monthly']) >= 2:
        processed_data['monthly'] = [data['monthly'][-2]]
        period_label = f"지난달({processed_data['monthly'][0]['month']})"

    if period_label:
        message += f"📅 {period_label} 생산 데이터:\n\n"

        for m in processed_data['monthly']:
            message += f"[{m['month']}]\n"
            message += f"🏭 총 생산량: {format_number(m['totalProduction'])}개\n"
            message += f"✅ 양품: {format_number(m['goodProducts'])}개\n"
            message += f"❌ 불량품: {format_number(m['defectProducts'])}개\n"
            message += f"📊 불량률: {m['defectRate']}%\n"
            message += f"⚡ 가동률: {m['efficiency']}%\n\n"

    if is_comparison and processed_data.get('monthly') and len(processed_data['monthly']) >= 2:
        first = processed_data['monthly'][0]
        last = processed_data['monthly'][-1]
        prod_diff = last['totalProduction'] - first['totalProduction']
        eff_diff = last['efficiency'] - first['efficiency']

        message += f"\n📊 비교 분석 ({first['month']} vs {last['month']}):\n"
        message += f"- 생산량 변화: {format_number(abs(prod_diff))}개 {'📈 증가' if prod_diff > 0 else '📉 감소'} ({(prod_diff / first['totalProduction'] * 100):.1f}%)\n"
        message += f"- 가동률 변화: {abs(eff_diff):.1f}% {'📈 향상' if eff_diff > 0 else '📉 하락'}\n"
        message += f"- 불량률: {first['defectRate']}% → {last['defectRate']}% ({'⚠️ 증가' if last['defectRate'] > first['defectRate'] else '✅ 감소'})\n"

    if specific_item:
        processed_data['products'] = [p for p in data['products'] if p['name'] == specific_item]
        if processed_data['products']:
            product = processed_data['products'][0]
            message += f"\n🏷️ {specific_item} 상세 정보:\n"
            message += f"- 생산량: {format_number(product['quantity'])}개\n"
            message += f"- 목표량: {format_number(product['target'])}개\n"
            message += f"- 달성률: {product['rate']}%\n"
            message += f"- 목표 달성까지: {format_number(product['target'] - product['quantity'])}개 남음\n"

            if product['rate'] >= 95:
                message += "\n✅ 목표 달성률이 우수합니다!"
            elif product['rate'] < 90:
                message += "\n⚠️ 목표 달성률이 낮습니다. 생산 증대가 필요합니다."

    if stats_type and not specific_item:
        quantities = [p['quantity'] for p in processed_data['products']]
        rates = [p['rate'] for p in processed_data['products']]

        if stats_type == 'max':
            max_qty = max(quantities)
            max_product = next(p for p in processed_data['products'] if p['quantity'] == max_qty)
            message += f"\n🏆 최다 생산 제품:\n"
            message += f"- {max_product['name']}: {format_number(max_product['quantity'])}개\n"
            message += f"- 목표 달성률: {max_product['rate']}%\n"
            processed_data['highlight'] = max_product

        elif stats_type == 'min':
            min_qty = min(quantities)
            min_product = next(p for p in processed_data['products'] if p['quantity'] == min_qty)
            message += f"\n⚠️ 최소 생산 제품:\n"
            message += f"- {min_product['name']}: {format_number(min_product['quantity'])}개\n"
            message += f"- 목표 달성률: {min_product['rate']}%\n"
            message += "- 생산 증대가 필요합니다.\n"
            processed_data['highlight'] = min_product

        elif stats_type == 'average':
            avg_qty = sum(quantities) / len(quantities)
            avg_rate = sum(rates) / len(rates)
            message += f"\n📊 평균 생산 데이터:\n"
            message += f"- 평균 생산량: {format_number(round(avg_qty))}개\n"
            message += f"- 평균 달성률: {avg_rate:.1f}%\n"

        elif stats_type == 'sum':
            total_qty = sum(quantities)
            message += f"\n💰 총 생산량: {format_number(total_qty)}개\n"

    if not message:
        message = '생산 현황을 조회했습니다.'

    return {
        'type': 'production',
        'data': processed_data,
        'message': message,
        'query': query
    }


# ---------- 구매 질문 처리 ----------
def handle_purchase_query(query, period, is_comparison, specific_item, stats_type):
    data = get_purchase_data()
    message = ''
    processed_data = copy.deepcopy(data)

    if specific_item:
        processed_data['orders'] = [o for o in processed_data['orders'] if specific_item in o['supplier']]
        if processed_data['orders']:
            total = sum(o['amount'] for o in processed_data['orders'])
            message += f"🏢 {specific_item} 발주 내역 ({len(processed_data['orders'])}건, 총 {format_number(total)}원):\n\n"
            for o in processed_data['orders']:
                message += f"- {o['item']}: {format_number(o['amount'])}원 ({o['status']})\n"
            message += "\n"
        else:
            message += f"🏢 {specific_item} 관련 발주 내역이 없습니다.\n\n"

    if '완료' in query:
        processed_data['orders'] = [o for o in processed_data['orders'] if o['status'] == '완료']
        message += f"✅ 완료된 발주: {len(processed_data['orders'])}건\n\n"
    elif '진행' in query or '대기' in query:
        processed_data['orders'] = [o for o in processed_data['orders'] if o['status'] != '완료']
        message += f"⏳ 진행중/대기 발주: {len(processed_data['orders'])}건\n\n"

    if stats_type and processed_data['orders']:
        amounts = [o['amount'] for o in processed_data['orders']]

        if stats_type == 'max':
            max_amount = max(amounts)
            max_order = next(o for o in processed_data['orders'] if o['amount'] == max_amount)
            message += f"\n💰 최대 발주:\n"
            message += f"- 공급업체: {max_order['supplier']}\n"
            message += f"- 품목: {max_order['item']}\n"
            message += f"- 금액: {format_number(max_amount)}원\n"
            message += f"- 상태: {max_order['status']}\n"
            processed_data['highlight'] = max_order

        elif stats_type == 'sum':
            total_amount = sum(amounts)
            message += f"\n📊 총 발주 금액: {format_number(total_amount)}원\n"
            message += f"- 발주 건수: {len(processed_data['orders'])}건\n"
            message += f"- 평균 발주액: {format_number(round(total_amount / len(processed_data['orders'])))}원\n"

        elif stats_type == 'average':
            avg_amount = sum(amounts) / len(amounts)
            message += f"\n📊 평균 발주 금액: {format_number(round(avg_amount))}원\n"

    if not message:
        message = '구매 현황을 조회했습니다.'

    return {
        'type': 'purchase',
        'data': processed_data,
        'message': message,
        'query': query
    }


# ---------- 품질 질문 처리 ----------
def handle_quality_query(query, period, is_comparison, specific_item, stats_type, is_cause_analysis):
    data = get_quality_data()
    message = ''
    processed_data = copy.deepcopy(data)

    if is_cause_analysis:
        message += "🔍 불량 원인 분석:\n\n"

        max_defect = max(data['defectTypes'], key=lambda d: d['count'])

        message += f"📊 주요 불량 유형: {max_defect['type']} ({max_defect['count']}건, {max_defect['rate']}%)\n\n"

        # 임계치 알림 (미니 플레이북): 특정 불량 유형이 전체 불량의
        # QUALITY_ALERT_THRESHOLD%를 넘으면 담당부서(품질부)에는 조치 필요 알림을,
        # 관련부서(생산부)에는 참고용 알림을 각각 생성한다.
        if max_defect['rate'] > QUALITY_ALERT_THRESHOLD:
            owner_dept, related_dept = '품질부', '생산부'
            summary = (
                f"{max_defect['type']} 비중이 {max_defect['rate']}%로 "
                f"임계치({QUALITY_ALERT_THRESHOLD}%)를 초과했습니다."
            )
            notify_if_new(
                owner_dept,
                f"⚠️ [조치필요 · 담당: {owner_dept}] {summary} 원인 분석 후 알림에서 조치 완료로 표시해주세요."
            )
            notify_if_new(
                related_dept,
                f"📋 [참고] {summary} 담당부서({owner_dept})에서 원인 분석 중입니다."
            )

        processed_data['causeAnalysis'] = {
            'mainDefect': max_defect['type'],
            'causes': []
        }

        if max_defect['type'] == '외관 불량':
            message += "💡 외관 불량 주요 원인:\n"
            message += "1. 작업자 숙련도 부족 (35%)\n"
            message += "2. 원자재 품질 문제 (28%)\n"
            message += "3. 설비 노후화 (22%)\n"
            message += "4. 작업 환경 (15%)\n\n"
            message += "✅ 개선 방안:\n"
            message += "- 작업자 교육 강화\n"
            message += "- 원자재 입고 검사 강화\n"
            message += "- 설비 정기 점검 및 교체\n"

            processed_data['causeAnalysis']['causes'] = [
                {'cause': '작업자 숙련도 부족', 'percent': 35},
                {'cause': '원자재 품질 문제', 'percent': 28},
                {'cause': '설비 노후화', 'percent': 22},
                {'cause': '작업 환경', 'percent': 15}
            ]
        elif max_defect['type'] == '치수 불량':
            message += "💡 치수 불량 주요 원인:\n"
            message += "1. 설비 캘리브레이션 오차 (42%)\n"
            message += "2. 온습도 변화 (28%)\n"
            message += "3. 측정 기구 오차 (20%)\n"
            message += "4. 원자재 변형 (10%)\n\n"
            message += "✅ 개선 방안:\n"
            message += "- 설비 정밀 캘리브레이션\n"
            message += "- 작업장 온습도 관리\n"
            message += "- 측정 기구 정기 교정\n"

            processed_data['causeAnalysis']['causes'] = [
                {'cause': '설비 캘리브레이션 오차', 'percent': 42},
                {'cause': '온습도 변화', 'percent': 28},
                {'cause': '측정 기구 오차', 'percent': 20},
                {'cause': '원자재 변형', 'percent': 10}
            ]
        elif max_defect['type'] == '기능 불량':
            message += "💡 기능 불량 주요 원인:\n"
            message += "1. 부품 조립 불량 (38%)\n"
            message += "2. 전기적 결함 (32%)\n"
            message += "3. 소프트웨어 오류 (20%)\n"
            message += "4. 부품 호환성 (10%)\n\n"
            message += "✅ 개선 방안:\n"
            message += "- 조립 공정 표준화\n"
            message += "- 전기 검사 강화\n"
            message += "- 소프트웨어 테스트 강화\n"

            processed_data['causeAnalysis']['causes'] = [
                {'cause': '부품 조립 불량', 'percent': 38},
                {'cause': '전기적 결함', 'percent': 32},
                {'cause': '소프트웨어 오류', 'percent': 20},
                {'cause': '부품 호환성', 'percent': 10}
            ]
        else:
            message += "💡 포장 불량 주요 원인:\n"
            message += "1. 포장 작업 미숙 (45%)\n"
            message += "2. 포장재 품질 (30%)\n"
            message += "3. 물류 과정 손상 (15%)\n"
            message += "4. 포장 설비 문제 (10%)\n\n"
            message += "✅ 개선 방안:\n"
            message += "- 포장 작업 교육\n"
            message += "- 고품질 포장재 사용\n"
            message += "- 물류 프로세스 개선\n"

            processed_data['causeAnalysis']['causes'] = [
                {'cause': '포장 작업 미숙', 'percent': 45},
                {'cause': '포장재 품질', 'percent': 30},
                {'cause': '물류 과정 손상', 'percent': 15},
                {'cause': '포장 설비 문제', 'percent': 10}
            ]

        return {
            'type': 'quality_cause',
            'data': processed_data,
            'message': message,
            'query': query
        }

    if specific_item:
        processed_data['defectTypes'] = [d for d in data['defectTypes'] if d['type'] == specific_item]
        if processed_data['defectTypes']:
            defect = processed_data['defectTypes'][0]
            message += f"📋 {specific_item} 분석:\n"
            message += f"- 발생 건수: {defect['count']}건\n"
            message += f"- 전체 불량 중 비율: {defect['rate']}%\n"
            message += f"- 총 검사 대비: {(defect['count'] / data['summary']['totalInspections'] * 100):.2f}%\n"

    if stats_type and not specific_item:
        counts = [d['count'] for d in processed_data['defectTypes']]

        if stats_type == 'max':
            max_count = max(counts)
            max_defect = next(d for d in processed_data['defectTypes'] if d['count'] == max_count)
            message += f"\n⚠️ 가장 많은 불량 유형:\n"
            message += f"- {max_defect['type']}: {max_defect['count']}건 ({max_defect['rate']}%)\n"
            message += "- ⚡ 우선 개선이 필요합니다!\n"
            processed_data['highlight'] = max_defect

        elif stats_type == 'min':
            min_count = min(counts)
            min_defect = next(d for d in processed_data['defectTypes'] if d['count'] == min_count)
            message += f"\n✅ 가장 적은 불량 유형:\n"
            message += f"- {min_defect['type']}: {min_defect['count']}건 ({min_defect['rate']}%)\n"
            message += "- 관리가 잘 되고 있습니다.\n"
            processed_data['highlight'] = min_defect

        elif stats_type == 'sum':
            total_defects = sum(counts)
            message += f"\n📊 총 불량 건수: {total_defects}건\n"
            message += f"- 전체 검사 대비: {(total_defects / data['summary']['totalInspections'] * 100):.2f}%\n"

    if '합격률' in query or '불량률' in query:
        message += f"\n📊 품질 지표:\n"
        message += f"- ✅ 합격률: {data['summary']['passRate']}%\n"
        message += f"- ❌ 불량률: {(100 - data['summary']['passRate']):.1f}%\n"
        message += f"- 📋 총 검사: {format_number(data['summary']['totalInspections'])}건\n"
        message += f"- ⚠️ 총 불량: {data['summary']['defectCount']}건\n"

    if not message:
        message = '품질 현황을 조회했습니다.'

    return {
        'type': 'quality',
        'data': processed_data,
        'message': message,
        'query': query
    }


# ---------- 인사 질문 처리 ----------
def handle_hr_query(query, period, is_comparison, specific_item, stats_type):
    data = get_hr_data()
    message = ''
    processed_data = copy.deepcopy(data)

    if specific_item:
        processed_data['departments'] = [d for d in data['departments'] if d['name'] == specific_item]
        if processed_data['departments']:
            dept = processed_data['departments'][0]
            message += f"👥 {specific_item} 상세 정보:\n"
            message += f"- 인원: {dept['employees']}명\n"
            message += f"- 출근율: {dept['attendance']}%\n"
            message += f"- 전체 직원 대비: {(dept['employees'] / data['summary']['totalEmployees'] * 100):.1f}%\n"

            if dept['attendance'] >= 97:
                message += "\n✅ 출근율이 우수합니다!"
            elif dept['attendance'] < 96:
                message += "\n⚠️ 출근율 관리가 필요합니다."

    if stats_type and not specific_item:
        employees = [d['employees'] for d in processed_data['departments']]
        attendances = [d['attendance'] for d in processed_data['departments']]

        if stats_type == 'max':
            if '인원' in query:
                max_emp = max(employees)
                max_dept = next(d for d in processed_data['departments'] if d['employees'] == max_emp)
                message += f"\n👥 최대 인원 부서:\n"
                message += f"- {max_dept['name']}: {max_dept['employees']}명\n"
                message += f"- 전체의 {(max_dept['employees'] / data['summary']['totalEmployees'] * 100):.1f}%\n"
                processed_data['highlight'] = max_dept
            else:
                max_att = max(attendances)
                max_dept = next(d for d in processed_data['departments'] if d['attendance'] == max_att)
                message += f"\n⭐ 최고 출근율 부서:\n"
                message += f"- {max_dept['name']}: {max_dept['attendance']}%\n"
                message += f"- 인원: {max_dept['employees']}명\n"
                processed_data['highlight'] = max_dept

        elif stats_type == 'min':
            if '인원' in query:
                min_emp = min(employees)
                min_dept = next(d for d in processed_data['departments'] if d['employees'] == min_emp)
                message += f"\n👥 최소 인원 부서:\n"
                message += f"- {min_dept['name']}: {min_dept['employees']}명\n"
                processed_data['highlight'] = min_dept
            else:
                min_att = min(attendances)
                min_dept = next(d for d in processed_data['departments'] if d['attendance'] == min_att)
                message += f"\n⚠️ 최저 출근율 부서:\n"
                message += f"- {min_dept['name']}: {min_dept['attendance']}%\n"
                message += "- 관리가 필요합니다.\n"
                processed_data['highlight'] = min_dept

        elif stats_type == 'average':
            avg_emp = sum(employees) / len(employees)
            avg_att = sum(attendances) / len(attendances)
            message += f"\n📊 부서별 평균:\n"
            message += f"- 평균 인원: {round(avg_emp)}명\n"
            message += f"- 평균 출근율: {avg_att:.1f}%\n"

        elif stats_type == 'sum':
            total_emp = sum(employees)
            message += f"\n👥 총 직원 수: {total_emp}명\n"

    if '입사' in query or '퇴사' in query or '이직' in query:
        message += f"\n📋 인력 변동 현황:\n"
        message += f"- 신규 입사: {data['summary']['newHires']}명\n"
        message += f"- 퇴사자: {data['summary']['resignations']}명\n"
        net_change = data['summary']['newHires'] - data['summary']['resignations']
        message += f"- 순증감: {'+' if net_change > 0 else ''}{net_change}명\n"
        message += f"- 이직률: {(data['summary']['resignations'] / data['summary']['totalEmployees'] * 100):.2f}%\n"

    if not message:
        message = '인사 현황을 조회했습니다.'

    return {
        'type': 'hr',
        'data': processed_data,
        'message': message,
        'query': query
    }


# ---------- 종합 질문 처리 ----------
DOMAIN_LABELS = {
    'financial': '재무', 'production': '생산', 'purchase': '구매',
    'quality': '품질', 'quality_cause': '품질(원인분석)', 'hr': '인사'
}


def get_domain_data(domain_type):
    if domain_type == 'financial':
        return get_financial_data()
    if domain_type == 'production':
        return get_production_data()
    if domain_type == 'purchase':
        return get_purchase_data()
    if domain_type in ('quality', 'quality_cause'):
        return get_quality_data()
    if domain_type == 'hr':
        return get_hr_data()
    return None


# ---------- ERP 데이터 기반 이슈 체크리스트 ----------
# RAG는 "사내 규정 문서"를 근거로 답하는 반면, 이건 실제 조회한 ERP 수치를
# 근거로 "지금 우선적으로 확인/조치해야 할 것"을 AI가 정리하도록 한다.
def generate_issue_checklist(domain_type, data):
    if not data:
        return None
    label = DOMAIN_LABELS.get(domain_type, domain_type)
    # 일별 원시 로그(daily/inspections/attendance)는 체크리스트 판단에 크게
    # 안 쓰이면서 프롬프트만 길어지게 하므로 제외해 응답 속도를 개선한다.
    trimmed = {k: v for k, v in data.items() if k not in ('daily', 'inspections', 'attendance')}
    data_text = json.dumps(trimmed, ensure_ascii=False, indent=2)
    prompt = (
        f'다음은 사내 ERP {label} 데이터입니다. 이 데이터를 분석해서 담당자가 지금 우선적으로 '
        '확인하거나 조치해야 할 사항을 중요도 순으로 번호를 매긴 체크리스트로 정리하세요. '
        '각 항목은 데이터의 구체적인 수치를 근거로 제시하고, 왜 우선순위가 높은지 한 줄로 '
        '설명하세요. 데이터에 없는 내용은 추측하지 마세요.\n\n'
        f'[{label} 데이터]\n{data_text}'
    )
    try:
        return call_potens_ai(prompt)
    except Exception as e:
        print(f'이슈 체크리스트 생성 실패: {e}')
        return None


def handle_comprehensive_query(query, allowed_categories=None):
    all_data = {
        'financial': get_financial_data()['summary'],
        'production': get_production_data()['summary'],
        'purchase': get_purchase_data()['summary'],
        'quality': get_quality_data()['summary'],
        'hr': get_hr_data()['summary']
    }

    # RBAC: '종합' 응답이라도 소속 부서 권한 밖 영역의 데이터는 포함하지 않는다.
    if allowed_categories is not None:
        visible_data = {k: v for k, v in all_data.items() if k in allowed_categories}
    else:
        visible_data = all_data

    if not visible_data:
        message = '🔒 소속 부서 권한으로 조회 가능한 종합 현황 항목이 없습니다.'
    else:
        labels = ', '.join(DOMAIN_LABELS[k] for k in visible_data)
        message = f'전체 경영 현황을 조회했습니다. (조회 가능 영역: {labels})'

    return {
        'type': 'comprehensive',
        'data': visible_data,
        'message': message,
        'query': query
    }


# ---------- 숫자 포맷팅 ----------
def format_number(num):
    return f"{num:,}"


# ---------- 재무 데이터 ----------
def get_financial_data():
    return {
        'summary': {
            'totalRevenue': 15800000000,
            'totalExpense': 12300000000,
            'netProfit': 3500000000,
            'profitRate': 22.15
        },
        'monthly': [
            {'month': '1월', 'revenue': 1200000000, 'expense': 950000000, 'profit': 250000000},
            {'month': '2월', 'revenue': 1350000000, 'expense': 1050000000, 'profit': 300000000},
            {'month': '3월', 'revenue': 1280000000, 'expense': 1020000000, 'profit': 260000000},
            {'month': '4월', 'revenue': 1420000000, 'expense': 1100000000, 'profit': 320000000},
            {'month': '5월', 'revenue': 1380000000, 'expense': 1080000000, 'profit': 300000000},
            {'month': '6월', 'revenue': 1470000000, 'expense': 1150000000, 'profit': 320000000}
        ]
    }


# ---------- 생산 데이터 ----------
def get_production_data():
    return {
        'summary': {
            'totalProduction': 125000,
            'defectRate': 1.8,
            'efficiency': 94.5
        },
        'monthly': [
            {'month': '1월', 'totalProduction': 19500, 'goodProducts': 19140, 'defectProducts': 360, 'defectRate': 1.85, 'efficiency': 93.2},
            {'month': '2월', 'totalProduction': 20200, 'goodProducts': 19838, 'defectProducts': 362, 'defectRate': 1.79, 'efficiency': 94.1},
            {'month': '3월', 'totalProduction': 20800, 'goodProducts': 20426, 'defectProducts': 374, 'defectRate': 1.80, 'efficiency': 94.8},
            {'month': '4월', 'totalProduction': 21500, 'goodProducts': 21115, 'defectProducts': 385, 'defectRate': 1.79, 'efficiency': 95.2},
            {'month': '5월', 'totalProduction': 21200, 'goodProducts': 20828, 'defectProducts': 372, 'defectRate': 1.75, 'efficiency': 94.6},
            {'month': '6월', 'totalProduction': 21800, 'goodProducts': 21404, 'defectProducts': 396, 'defectRate': 1.82, 'efficiency': 95.4}
        ],
        'products': [
            {'name': 'A-100 모델', 'quantity': 35000, 'target': 40000, 'rate': 87.5},
            {'name': 'B-200 모델', 'quantity': 42000, 'target': 45000, 'rate': 93.3},
            {'name': 'C-300 모델', 'quantity': 28000, 'target': 30000, 'rate': 93.3},
            {'name': 'D-400 모델', 'quantity': 20000, 'target': 22000, 'rate': 90.9}
        ],
        'daily': [
            {'date': '2026-07-10', 'production': 2100, 'defect': 38},
            {'date': '2026-07-11', 'production': 2250, 'defect': 41},
            {'date': '2026-07-12', 'production': 2180, 'defect': 35},
            {'date': '2026-07-13', 'production': 2300, 'defect': 42},
            {'date': '2026-07-14', 'production': 2150, 'defect': 39}
        ]
    }


# ---------- 구매 데이터 ----------
def get_purchase_data():
    return {
        'summary': {
            'totalOrders': 245,
            'totalAmount': 4800000000,
            'pendingOrders': 18
        },
        'orders': [
            {'supplier': '(주)대한소재', 'item': '철강 원자재', 'amount': 850000000, 'status': '완료'},
            {'supplier': '글로벌부품', 'item': '전자부품 세트', 'amount': 620000000, 'status': '진행중'},
            {'supplier': '한국화학', 'item': '산업용 화학제', 'amount': 450000000, 'status': '완료'},
            {'supplier': '프리미엄자재', 'item': '특수 합금', 'amount': 720000000, 'status': '대기'},
            {'supplier': '스마트부품상사', 'item': 'PCB 기판', 'amount': 380000000, 'status': '완료'}
        ]
    }


# ---------- 품질 데이터 ----------
def get_quality_data():
    return {
        'summary': {
            'totalInspections': 8500,
            'passRate': 98.2,
            'defectCount': 153
        },
        'defectTypes': [
            {'type': '외관 불량', 'count': 62, 'rate': 40.5},
            {'type': '치수 불량', 'count': 38, 'rate': 24.8},
            {'type': '기능 불량', 'count': 28, 'rate': 18.3},
            {'type': '포장 불량', 'count': 25, 'rate': 16.4}
        ],
        'inspections': [
            {'date': '2026-07-10', 'total': 1700, 'pass': 1668, 'fail': 32},
            {'date': '2026-07-11', 'total': 1750, 'pass': 1719, 'fail': 31},
            {'date': '2026-07-12', 'total': 1680, 'pass': 1651, 'fail': 29},
            {'date': '2026-07-13', 'total': 1800, 'pass': 1767, 'fail': 33},
            {'date': '2026-07-14', 'total': 1570, 'pass': 1542, 'fail': 28}
        ]
    }


# ---------- 인사 데이터 ----------
def get_hr_data():
    return {
        'summary': {
            'totalEmployees': 485,
            'newHires': 12,
            'resignations': 5,
            'avgAttendance': 96.8
        },
        'departments': [
            {'name': '생산부', 'employees': 180, 'attendance': 97.2},
            {'name': '영업부', 'employees': 85, 'attendance': 95.8},
            {'name': '기술부', 'employees': 95, 'attendance': 96.5},
            {'name': '관리부', 'employees': 65, 'attendance': 97.8},
            {'name': '품질부', 'employees': 60, 'attendance': 98.1}
        ],
        'attendance': [
            {'date': '2026-07-10', 'present': 472, 'absent': 8, 'leave': 5},
            {'date': '2026-07-11', 'present': 468, 'absent': 10, 'leave': 7},
            {'date': '2026-07-12', 'present': 475, 'absent': 6, 'leave': 4},
            {'date': '2026-07-13', 'present': 470, 'absent': 9, 'leave': 6},
            {'date': '2026-07-14', 'present': 473, 'absent': 7, 'leave': 5}
        ]
    }


# ---------- 카테고리별 데이터 조회 ----------
@app.route('/api/category/<category>', methods=['GET'])
@login_required
def get_category_data_route(category):
    allowed = get_allowed_categories(session['department'], session['role'])
    if category not in allowed:
        result = {
            'type': 'text',
            'data': None,
            'message': '🔒 소속 부서 권한으로는 조회할 수 없는 데이터입니다. 담당 부서 또는 관리자에게 문의해주세요.'
        }
    else:
        # /api/query와 동일하게 신뢰도 표시·관련 메뉴 안내를 붙이고 조회 이력에도 남긴다
        # (사이드바 '빠른 조회' 버튼과 채팅 질의가 서로 다른 응답을 주지 않도록 통일).
        result = enrich_structured_response(get_category_data(category))

    conversation_id = request.args.get('conversation_id', type=int)
    label = DOMAIN_LABELS.get(category, category)
    if not conversation_id:
        conversation_id = create_conversation(session['user_id'], f'[빠른 조회] {label}')
    else:
        touch_conversation(conversation_id)

    log_query_history(
        session['user_id'], conversation_id, f'[빠른 조회] {label}',
        result.get('type', 'text'), result.get('message', ''), result.get('data')
    )
    result['conversation_id'] = conversation_id
    return jsonify(result)


# ---------- 알림 조회/읽음 처리 ----------
@app.route('/api/notifications', methods=['GET'])
@login_required
def notifications_route():
    unread_only = request.args.get('unread') == '1'
    items = get_notifications(session['department'], session['role'], unread_only)
    return jsonify({'notifications': items})


@app.route('/api/notifications/<int:notification_id>/read', methods=['POST'])
@login_required
def mark_notification_read_route(notification_id):
    comment = (request.json or {}).get('comment') if request.is_json else None

    conn = get_db()
    if comment:
        conn.execute(
            'UPDATE notifications SET is_read = 1, resolution_comment = ? WHERE id = ?',
            (comment, notification_id)
        )
    else:
        conn.execute('UPDATE notifications SET is_read = 1 WHERE id = ?', (notification_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


# ---------- 대화 세션 라우트 ----------
@app.route('/api/conversations', methods=['GET'])
@login_required
def conversations_route():
    return jsonify({'conversations': get_conversations(session['user_id'])})


@app.route('/api/conversations/<int:conversation_id>/messages', methods=['GET'])
@login_required
def conversation_messages_route(conversation_id):
    return jsonify({'messages': get_conversation_messages(conversation_id, session['user_id'])})


@app.route('/api/conversations/<int:conversation_id>/favorite', methods=['POST'])
@login_required
def toggle_conversation_favorite_route(conversation_id):
    conn = get_db()
    row = conn.execute(
        'SELECT is_favorite FROM conversations WHERE id = ? AND user_id = ?',
        (conversation_id, session['user_id'])
    ).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': '대화를 찾을 수 없습니다.'}), 404

    new_value = 0 if row['is_favorite'] else 1
    conn.execute('UPDATE conversations SET is_favorite = ? WHERE id = ?', (new_value, conversation_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'is_favorite': bool(new_value)})


@app.route('/api/generate-faq', methods=['POST'])
@login_required
def generate_faq_route():
    if session['role'] != 'admin':
        return jsonify({'error': '관리자만 사용할 수 있습니다.'}), 403

    path = generate_faq_document()
    if not path:
        return jsonify({'ok': False, 'message': '아직 반복 질의가 충분히 쌓이지 않았습니다 (동일 질문 2회 이상 필요).'})
    return jsonify({'ok': True, 'message': 'FAQ 문서를 생성해 사내 지식 문서(knowledge)에 반영했습니다.'})


@app.route('/api/generate-department-guide', methods=['POST'])
@login_required
def generate_department_guide_route():
    department = session['department']
    allowed = get_allowed_categories(department, session['role'])
    guide = build_department_guide(department, allowed)
    if not guide:
        return jsonify({'ok': False, 'message': '조회 권한이 없어 가이드를 생성할 수 없습니다.'})
    save_department_guide_to_knowledge(department, guide['markdown'])
    return jsonify({'ok': True, 'department': department, 'content': guide['markdown']})


@app.route('/api/generate-department-guide/excel', methods=['POST'])
@login_required
def generate_department_guide_excel_route():
    department = session['department']
    allowed = get_allowed_categories(department, session['role'])
    guide = build_department_guide(department, allowed)
    if not guide or not guide['rows']:
        return jsonify({'error': '내보낼 가이드 내용이 없습니다.'}), 400
    save_department_guide_to_knowledge(department, guide['markdown'])
    return build_excel_response(
        'department_guide', ['분류', '관련메뉴', '예시질문'], guide['rows'],
        filename=f'{department}_ERP가이드.xlsx'
    )


# ---------- 응답 CSV 내보내기 ----------
def extract_export_rows(export_type, data):
    if not data:
        return None, None
    if export_type == 'financial' and data.get('monthly'):
        return ['month', 'revenue', 'expense', 'profit'], data['monthly']
    if export_type == 'production':
        if data.get('monthly'):
            return ['month', 'totalProduction', 'goodProducts', 'defectProducts', 'defectRate', 'efficiency'], data['monthly']
        if data.get('products'):
            return ['name', 'quantity', 'target', 'rate'], data['products']
    if export_type == 'purchase' and data.get('orders'):
        return ['supplier', 'item', 'amount', 'status'], data['orders']
    if export_type == 'quality' and data.get('defectTypes'):
        return ['type', 'count', 'rate'], data['defectTypes']
    if export_type == 'quality_cause' and data.get('causeAnalysis', {}).get('causes'):
        return ['cause', 'percent'], data['causeAnalysis']['causes']
    if export_type == 'hr' and data.get('departments'):
        return ['name', 'employees', 'attendance'], data['departments']
    return None, None


def content_disposition_header(filename):
    # HTTP 헤더 값은 latin-1만 허용되어 한글 파일명을 그대로 넣으면 응답 자체가
    # 깨진다(서버가 응답을 못 보내고 클라이언트가 무한정 기다리게 됨). ASCII
    # 대체 파일명 + RFC 5987 형식의 UTF-8 파일명을 함께 내려준다.
    ascii_fallback = filename.encode('ascii', 'ignore').decode('ascii').strip() or 'download'
    return f"attachment; filename=\"{ascii_fallback}\"; filename*=UTF-8''{quote(filename)}"


EXPORT_SHEET_TITLES = {
    'financial': '재무',
    'production': '생산',
    'purchase': '구매',
    'quality': '품질',
    'quality_cause': '불량원인',
    'hr': '인사'
}


def build_excel_response(export_type, fieldnames, rows, filename=None):
    wb = Workbook()
    ws = wb.active
    ws.title = (EXPORT_SHEET_TITLES.get(export_type, export_type) or export_type)[:31]

    header_fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    ws.append(fieldnames)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        ws.append([row.get(f, '') for f in fieldnames])

    for i, col in enumerate(fieldnames, start=1):
        max_len = max([len(str(col))] + [len(str(row.get(col, ''))) for row in rows])
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max_len + 4

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    final_filename = filename or f'erp_export_{export_type}.xlsx'
    return Response(
        buffer.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': content_disposition_header(final_filename)}
    )


@app.route('/api/export', methods=['POST'])
@login_required
def export_route():
    payload = request.json or {}
    export_type = payload.get('type')
    data = payload.get('data')
    export_format = payload.get('format', 'excel')

    fieldnames, rows = extract_export_rows(export_type, data)
    if not rows:
        return jsonify({'error': '내보낼 표 형식 데이터가 없습니다.'}), 400

    if export_format == 'excel':
        return build_excel_response(export_type, fieldnames, rows)

    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=fieldnames, extrasaction='ignore')
    writer.writeheader()
    writer.writerows(rows)

    # Excel에서 한글이 깨지지 않도록 BOM 포함 UTF-8로 인코딩
    csv_bytes = buffer.getvalue().encode('utf-8-sig')
    return Response(
        csv_bytes,
        mimetype='text/csv',
        headers={'Content-Disposition': content_disposition_header(f'erp_export_{export_type}.csv')}
    )


def get_category_data(category):
    if category == 'financial':
        return {'type': 'financial', 'data': get_financial_data(), 'message': '재무 현황입니다.'}
    if category == 'production':
        return {'type': 'production', 'data': get_production_data(), 'message': '생산 현황입니다.'}
    if category == 'purchase':
        return {'type': 'purchase', 'data': get_purchase_data(), 'message': '구매 현황입니다.'}
    if category == 'quality':
        return {'type': 'quality', 'data': get_quality_data(), 'message': '품질 현황입니다.'}
    if category == 'hr':
        return {'type': 'hr', 'data': get_hr_data(), 'message': '인사 현황입니다.'}
    return {'type': 'text', 'data': None, 'message': '데이터를 찾을 수 없습니다.'}


init_db()

if __name__ == '__main__':
    # threaded=True 필수: 기본(싱글 스레드) 상태에서는 한 사용자가 AI 응답을
    # 기다리는 동안 다른 모든 사용자의 요청이 전부 멈춰버린다.
    app.run(debug=True, port=int(os.environ.get('PORT', 5000)), threaded=True)
